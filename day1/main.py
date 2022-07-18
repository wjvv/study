import openpyxl
# 创建一个工作簿
# wb = openpyxl.Workbook()
# 创建一个test_case的sheet表单
# wb.create_sheet('test_case')
# 保存为一个xlsx格式的文件
# wb.save('C:\\Users\\Administrator\\Desktop\\cases1.xlsx')
# 读取excel中的数据
# 第一步：打开工作簿
wb = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\excel.xlsx')
# 第二步：选取表单
sh = wb['工作表格1']
# 第三步：读取数据
# 参数 row:行  column：列
ce = sh.cell(row=1, column=1)   # 读取第一行，第一列的数据
print(ce.value)
# 按行读取数据 list(sh.rows)
print(list(sh.rows)[1:])     # 按行读取数据，去掉第一行的表头信息数据
for cases in list(sh.rows)[1:]:
    case_id = cases[0].value
    case_excepted = cases[1].value
    case_data = cases[2].value
    print(case_id, case_excepted, case_data)
# 关闭工作薄
wb.close()

'''
import matplotlib.pyplot as plt

from openpyxl import workbook, load_workbook

wb = load_workbook('C:\\Users\\Administrator\\Desktop\\excel.xlsx')
ws = wb.active
print(ws['B3'].value)
ws['B3'].value = 'haha'
wb.save('C:\\Users\\Administrator\\Desktop\\excel.xlsx')
print(wb.sheetnames)

x = [1, 2, 3]
y = [1, 2, 3]
# 以下两种写法等价，
plt.plot(x, y, color='green', marker='+', linestyle='dashed', linewidth=2, markersize=12)
# plt.flot(x, y, 'go--'，linewidth=2, markersize=12)
# 可以在一个画布上绘制多张图片，
y1 = [4, 5, 6]
plt.plot(x, y1, color='red', marker='*', linestyle='solid', linewidth=2, markersize=12)
plt.show()
'''

